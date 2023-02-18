#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct  4 21:35:35 2022

@author: wae16sru
"""


# this file does t1 fitting according to SEPAL 
# signal to concentration conversion was done by SEPAL too 
#for drift calcualtion we are following the process of Heye et al 2016
#" Signal drift was calculated as the
#overall change in signal intensity per minute and is given as percentage
#of the time-averaged signal"
#we did both enhancement and intensity calcualtion to check 
# wehther there is any differece in the drift 
import os

sub_ID= input("enter scanner drift subject ID (e.g. MRIPD_052): ")


results_dir= "/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_manuscript_results/BBB_drift_results" 
dir= "/gpfs/home/wae16sru/BBB_sample_data/Scanner_drift_study/Scanner_drift_data/NIFTI_data/" + sub_ID
os.chdir(str(dir))


import sys
import numpy as np
import matplotlib.pyplot as plt
sys.path.append('/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/SEPAL-master/src')
import t1_fit,dce_fit, relaxivity, signal_models, water_ex_models, aifs, pk_models

#import from Donnie

import math
import numpy as np
import nibabel as nib
import matplotlib.pyplot as plt
plt.style.use( 'ggplot' )
plt.ion()
from scipy.optimize import minimize, newton, basinhopping, least_squares

#import for excel I/O

import sys
# import os
import openpyxl
import pandas as pd
from openpyxl import Workbook

sys.path.append('/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification')



## MAIN CODE LOOP_from DONNIE
# Load spoiled GRE images (in nifti format) from the working directory
T1img1 = nib.load('despot1_5_BET_MSK.nii.gz')
T1img1_data = T1img1.get_data( )
hdr_T1img1 = T1img1.header

T1img2 = nib.load('despot1_10_BET_MSK.nii.gz')
T1img2_data = T1img2.get_data( )

T1img3 = nib.load('despot1_15_BET_MSK.nii.gz')
T1img3_data = T1img3.get_data( )

B1img = nib.load('despot1_IR5_BET_MSK.nii.gz')
B1img_data = B1img.get_data( )

# Load DCE time series data from working directory
DCEimg = nib.load('dce_BET_MSK.nii.gz')
DCEimg_data = DCEimg.get_data( )

# Load parenchyma/CSF segmentation masks from working directory
#WM = nib.load('t1S_seg_2.nii.gz') # White matter
#WM_mask = WM.get_data( )
#GM = nib.load('t1S_seg_1.nii.gz') # Grey matter
#GM_mask = GM.get_data( )
#CSF = nib.load('t1S_seg_0.nii.gz') # Cerebrospinal fluid
#CSF_mask = CSF.get_data( )

WM= nib.load (dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/total_WM.nii.gz")
WM_mask = WM.get_data( )
CSF= nib.load (dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/CSF_fastsurfer_rawavg.nii.gz")
CSF_mask = CSF.get_data( )
GM= nib.load (dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/total_GM.nii.gz")
GM_mask = GM.get_data( )
hippo= nib.load(dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/total_hippo.nii.gz")
hippo_mask=hippo.get_data()
thalamus= nib.load(dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/total_thalamus.nii.gz")
thalamus_mask= thalamus.get_data()
amygdala= nib.load(dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/total_amygdala.nii.gz")
amygdala_mask= amygdala.get_data()
whole_brain=nib.load(dir + "/brain_mask_FINAL.nii.gz")
whole_brain_mask=whole_brain.get_data()

# TIME SERIES SIGNAL ENHANCEMENT CALCULATIONS
#==============================================
# Calculate timings of each acquisition time point. ** USER INPUT **
acq_len = 591.5  # Acquisition duration (s) (subtracted 31.5 so pt 0 = time 0)
nPts = len( DCEimg_data[ 1, 1, 1, : ] )
tVec = np.arange( 0, nPts ) * acq_len / nPts


# The vascular input function is calculated from a voxel in the superior
# sagittal sinus. Give co-ordinates for that voxel here.

#ss_x= input("enter the VIF x: ")
#ss_y= input("enter the VIF y: ")
#ss_z= input("enter the VIF z: ")

#ss_vox= [int("ss_x",base=10), int("ss_y",base=10), int("ss_z",base=10)]
#ss_vox= [ss_x,ss_y,ss_z]

# Calculate median signal intensities for each tissue type, for each time point
# Initialise arrays to hold median tissue SIs per image
med_WM = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )  # No baseline
med_GM = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )
med_CSF = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )
med_hippo = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )
med_thalamus = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )
med_amygdala = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )
# No median measurement for sagittal sinus - we use one voxel only
#SS = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )

# No median signal intensity at baseline from DESPOT1_15deg ( different TR/TE )

# Calculate median signal per time point by looping through DCE images
for i in range( 0, len( DCEimg_data[ 1, 1, 1, : ] ) ):
    DCEtPt = DCEimg_data[ :, :, :, i ]  # Split DCE data into time points
    med_WM[ i ] = np.median( DCEtPt[ np.where( WM_mask > 0 ) ] )
    med_GM[ i ] = np.median( DCEtPt[ np.where(( WM_mask == 0) & ( CSF_mask == 0) & (whole_brain_mask > 0)) ] )
    med_CSF[ i ] = np.median( DCEtPt[ np.where( CSF_mask > 0 ) ] )
    med_hippo [i]= np.median( DCEtPt[ np.where( hippo_mask > 0 ) ] )
    med_thalamus [i]= np.median( DCEtPt[ np.where( thalamus_mask > 0 ) ] )
    med_amygdala [i]= np.median( DCEtPt[ np.where( amygdala_mask > 0 ) ] )
    #SS[ i ] = DCEtPt[ ss_vox[ 0 ], ss_vox[ 1 ], ss_vox[ 2 ] ]

#normalise the signal INTENSITIE
sigIn_WM_norm=med_WM/med_WM.mean(0)
sigIn_GM_norm=med_GM/med_GM.mean(0)
sigIn_CSF_norm=med_CSF/med_CSF.mean(0)
sigIn_hippo_norm=med_hippo/med_hippo.mean(0)
sigIn_thalamus_norm=med_thalamus/med_thalamus.mean(0)
sigIn_amygdala_norm=med_amygdala/med_amygdala.mean(0)



# Convert to signal enhancement from baseline (ratio, use % for plotting only)
sigEn_WM = ( med_WM - med_WM[ 0 ] ) / med_WM[ 0 ]
sigEn_GM = ( med_GM - med_GM[ 0 ] ) / med_GM[ 0 ]
sigEn_CSF = ( med_CSF - med_CSF[ 0 ] ) / med_CSF[ 0 ]
sigEn_hippo = ( med_hippo - med_hippo[ 0 ] ) / med_hippo[ 0 ]
sigEn_thalamus = ( med_thalamus - med_thalamus[ 0 ] ) / med_thalamus[ 0 ]
sigEn_amygdala = ( med_amygdala - med_amygdala[ 0 ] ) / med_amygdala[ 0 ]
#????more regions to add here 
#sigEn_SS = ( SS - SS[ 0 ] ) / SS[ 0 ]


sigEn_tissue=[sigEn_WM, sigEn_GM, sigEn_CSF, sigEn_hippo, sigEn_thalamus, sigEn_amygdala] # ???need modification after more tissue regions 
# changes to adapt SEPAL 
sigIn_tissue= [med_WM, med_GM, med_CSF, med_hippo, med_thalamus, med_amygdala]

sigIn_tissue_norm=[sigIn_WM_norm, sigIn_GM_norm, sigIn_CSF_norm, sigIn_hippo_norm, sigIn_thalamus_norm, sigIn_amygdala_norm]


'''DCam. The first few participants have different TR/TE for DCE-MRI. Will use
t = 0 DCE timepoint as baseline instead ...'''

# Plotting code - shows signal enahncement time-curves for WM, GM, CSF, sigEn_SS SS
fig = plt.figure( )
fig.suptitle( f"Drift_{sub_ID} Enhancement", fontsize = 14 )
ax = plt.gca( )
plt1, = plt.plot( tVec, sigEn_WM * 100, '#56B4E9', label = 'White Matter', \
linewidth = 2 )
plt2, = plt.plot( tVec, sigEn_GM * 100, '#0072B2', label = 'Grey Matter' , \
linewidth = 2 )
plt3, = plt.plot( tVec, sigEn_hippo * 100, '#CC79A7', label = 'Hippocampus' , \
linewidth = 2 )
plt4, = plt.plot( tVec, sigEn_thalamus * 100, '#00FFFF', label = 'Thalamus' , \
linewidth = 2 )
plt5, = plt.plot( tVec, sigEn_amygdala * 100, '#8A2BE2', label = 'Amygdala' , \
linewidth = 2 )
plt6, = plt.plot( tVec, sigEn_CSF * 100, '#8B7355', label ='CSF' , \
linewidth = 2 )
#????more regions to add here 

ax.set_xlabel( 'Time (s)' )
ax.set_ylabel( 'Signal Enhancement (%)' )
ax.grid(None)



plt.legend( handles = [ plt1, plt2, plt3, plt4, plt5, plt6], loc = 'lower center', bbox_to_anchor=(0.0, -0.35,1.0, 0.5), mode = "expand", ncol = 3,frameon = False ) 
#????more regions to add here np.sum (sigIn_GM_drift)/ np. mean (sigIn_GM_drift)
plt.show( )
plt.savefig(f"{results_dir}/sigEn_drift_plot_{sub_ID}.png", format = 'png', bbox_inches = 'tight', dpi = 500)




# Plotting code - shows normalised signal intensity time-curves for WM, GM, CSF, sigEn_SS SS
fig = plt.figure( )
fig.suptitle( f"Drift_{sub_ID} Normalised Intensity", fontsize = 14 )
ax = plt.gca( )
plt1, = plt.plot( tVec, sigIn_WM_norm * 100, '#56B4E9', label = 'White Matter', \
linewidth = 2 )
plt2, = plt.plot( tVec, sigIn_GM_norm * 100, '#0072B2', label = 'Grey Matter' , \
linewidth = 2 )
plt3, = plt.plot( tVec, sigIn_hippo_norm * 100, '#CC79A7', label = 'Hippocampus' , \
linewidth = 2 )
plt4, = plt.plot( tVec, sigIn_thalamus_norm * 100, '#00FFFF', label = 'Thalamus' , \
linewidth = 2 )
plt5, = plt.plot( tVec, sigIn_amygdala_norm * 100 , '#8A2BE2', label = 'Amygdala' , \
linewidth = 2 )
plt6, = plt.plot( tVec, sigIn_CSF_norm * 100, '#8B7355', label ='CSF' , \
linewidth = 2 )
#????more regions to add here 

ax.set_xlabel( 'Time (s)' )
ax.set_ylabel( ' Normalised Signal Instensity (%)' )
ax.grid(None)

# Put SS on similar scale




plt.legend( handles = [ plt1, plt2, plt3, plt4, plt5, plt6], loc = 'lower center', bbox_to_anchor=(0.0, -0.35,1.0, 0.5), mode = "expand", ncol = 3,frameon = False ) 
#????more regions to add here np.sum (sigIn_GM_drift)/ np. mean (sigIn_GM_drift)
plt.show( )
plt.savefig(results_dir+'/'+ 'sigIn_norm_drift_plot_'+f'{sub_ID}.png', format = 'png', bbox_inches = 'tight', dpi = 500)




medWM_sigSPGR = [ np.median( T1img1_data[ np.where( WM_mask > 0 ) ] ), \
np.median( T1img2_data[ np.where( WM_mask > 0 ) ] ), \
med_WM[ 0 ] ] # Arrange WM SPGR data and redefine med_WM[ 0 ] for clarity
medWM_sigIR = np.median( B1img_data[ np.where( WM_mask > 0 ) ] )

medGM_sigSPGR = [ np.median( T1img1_data[ np.where( ( WM_mask == 0) & ( CSF_mask == 0) & (whole_brain_mask > 0))   ] ), \
np.median( T1img2_data[ np.where( ( WM_mask == 0) & ( CSF_mask == 0) & (whole_brain_mask > 0) ) ] ), \
med_GM[ 0 ] ] # Arrange GM SPGR data and redefine med_GM[ 0 ] for clarity
medGM_sigIR = np.median( B1img_data[ np.where( ( WM_mask == 0) & ( CSF_mask == 0) & (whole_brain_mask > 0)) ] )

medCSF_sigSPGR = [ np.median( T1img1_data[ np.where( CSF_mask > 0 ) ] ), \
np.median( T1img2_data[ np.where( CSF_mask > 0 ) ] ), \
med_CSF[ 0 ] ] # Arrange CSF SPGR data and redefine med_CSF[ 0 ] for clarity
medCSF_sigIR = np.median( B1img_data[ np.where( CSF_mask > 0 ) ] )

#????more regions to add here 

medhippo_sigSPGR = [ np.median( T1img1_data[ np.where( hippo_mask > 0 ) ] ), \
np.median( T1img2_data[ np.where( hippo_mask > 0 ) ] ), \
med_hippo[ 0 ] ] # Arrange CSF SPGR data and redefine med_CSF[ 0 ] for clarity
medhippo_sigIR = np.median( B1img_data[ np.where( hippo_mask > 0 ) ] )


medthalamus_sigSPGR = [ np.median( T1img1_data[ np.where ( thalamus_mask > 0 ) ] ), \
np.median( T1img2_data[ np.where( thalamus_mask > 0 ) ] ), \
med_thalamus[ 0 ] ] # Arrange CSF SPGR data and redefine med_CSF[ 0 ] for clarity
medthalamus_sigIR = np.median( B1img_data[ np.where( thalamus_mask > 0 ) ] )

medamygdala_sigSPGR = [ np.median( T1img1_data[ np.where(amygdala_mask > 0 ) ] ), \
np.median( T1img2_data[ np.where(amygdala_mask > 0 ) ] ), \
med_amygdala[ 0 ] ] # Arrange CSF SPGR data and redefine med_CSF[ 0 ] for clarity
medamygdala_sigIR = np.median( B1img_data[ np.where( amygdala_mask > 0 ) ] )



#DESPOT_HIFI method from Michael (demo_fit_t1)


#book_result= Workbook()
#book_result= openpyxl.load_workbook('/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_scanner_drift.xlsx')
#book_result= Workbook()
#sheet_result=book_result.active
#sheet_result.cell(row=int(sub_ID[0])+1, column = 2*(tissue_type)+7).value= vp
#sheet_result.cell(row=int(sub_ID[0])+1, column= 2*(tissue_type)+8).value= ps

#book_result.save("/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_results.xlsx")

#df = pd.DataFrame(pd.read_excel('/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_results.xlsx'))
#df= pd.DataFrame(pd.read_excel("C:\\Users\\wae16sru\\OneDrive - University of East Anglia\\BBB_results\\BBB_results.xlsx"))

#print(df)
#entering signal enhancement into a csv file 
sigEn_tissue_df = pd.DataFrame (sigEn_tissue)

sigEn_tissue_df.index = ['WM','GM','CSF','Hippo','Thalamus','Amygdala']

sigEn_tissue_df.to_csv(results_dir+'/'+'scanner_drift_sigEn.csv', mode='a', header= False)

# same work for intensity 
sigIn_tissue_df = pd.DataFrame (sigIn_tissue)

sigIn_tissue_df.index = ['WM','GM','CSF','Hippo','Thalamus','Amygdala']

sigIn_tissue_df.to_csv(results_dir+'/'+ 'scanner_drift_sigIn.csv', mode='a', header= False)
 
# print message
print("Data saved successfully in subject directory")


# same work for noramlise intensity 

sigIn_tissue_norm_df = pd.DataFrame (sigIn_tissue_norm)

sigIn_tissue_norm_df.index = ['WM','GM','CSF','Hippo','Thalamus','Amygdala']

sigIn_tissue_norm_df.to_csv(results_dir+'/'+'scanner_drift_sigIn_norm.csv', mode='a', header= False)


# print message
print("Data saved successfully in subject's directory")

# same work for intensity as for enhancement below 
sigIn_tissue_allsub_df= pd.DataFrame (sigIn_tissue)

sigIn_tissue_allsub_df.index = ['WM','GM','CSF','Hippo','Thalamus','Amygdala']

sigIn_tissue_allsub_df.to_csv(results_dir+'/'+'BBB_drift_sigIn_allsub_on_T1_std.csv', mode='a', header= False)

print("Data appended successfully in directory: /gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_manuscript_results ")
 

sigEn_tissue_allsub_df= pd.DataFrame (sigEn_tissue)

sigEn_tissue_allsub_df.index = ['WM','GM','CSF','Hippo','Thalamus','Amygdala']

sigEn_tissue_allsub_df.to_csv(results_dir+'/'+'BBB_drift_sigEn_allsub_on_T1_std.csv', mode='a', header= False)

print("Data appended successfully in directory: /gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_manuscript_results ")

# same work for noramlised intensity as for enhancement 
sigIn_tissue_norm_allsub_df= pd.DataFrame (sigIn_tissue_norm)

sigIn_tissue_norm_allsub_df.index = ['WM','GM','CSF','Hippo','Thalamus','Amygdala']

sigIn_tissue_norm_allsub_df.to_csv(results_dir+'/'+'BBB_drift_sigIn_norm_allsub_on_T1_std.csv', mode='a', header= False)

print("Data appended successfully in directory: /gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_manuscript_results")

#this part fo teh code perform linear fitting as per Armitage 2011 to find drift 
# as per Armitage drift was calcualted by analyzing the slope of the signal
# enhancement profiles using standard linear regression
# analysis, we did linear fitting 
#As per armitage, he considered enhancement , so th

slope_intercept_En= np.zeros((6,2))
slope_intercept_In=np.zeros((6,2))
slope_intercept_In_norm=np.zeros((6,2))


for tissue_type in range(6):
    slope_intercept_In[tissue_type,:] = np.polyfit(tVec/60,sigIn_tissue[tissue_type],1)
    slope_intercept_In_norm[tissue_type,:] = np.polyfit(tVec/60,sigIn_tissue_norm[tissue_type],1)
    slope_intercept_En[tissue_type,:] = np.polyfit(tVec/60,sigEn_tissue[tissue_type],1)
    
drift_In_tissue=slope_intercept_In[:,0]
drift_En_tissue=slope_intercept_En[:,0]
drift_In_norm_tissue=slope_intercept_In_norm[:,0]


# ROI-wise drift for the selected subject , calucaltion on noramlised signal intensity 

drift_In_norm_tissue_df= pd.DataFrame (drift_In_norm_tissue)

drift_In_norm_tissue_df.index = ['WM','GM','CSF','Hippo','Thalamus','Amygdala']

drift_In_norm_tissue_df.to_csv(results_dir+'/'+'BBB_drift_polyfit_sigIn_norm_on_T1_std.csv', mode='a', header= False)

print("Data appended successfully in directory: /gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_manuscript_results")

# ROI-wise drift for the selected subject , calucaltion on  signal intensity 
drift_In_tissue_df= pd.DataFrame (drift_In_tissue)

drift_In_tissue_df.index = ['WM','GM','CSF','Hippo','Thalamus','Amygdala']

drift_In_tissue_df.to_csv(results_dir+'/'+'BBB_drift_polyfit_sigIn_on_T1_std.csv', mode='a', header= False)

print("Data appended successfully in directory: /gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_manuscript_results")


# ROI-wise drift for the selected subject , calucaltion on signal enhancement 
drift_En_tissue_df= pd.DataFrame (drift_En_tissue)

drift_En_tissue_df.index = ['WM','GM','CSF','Hippo','Thalamus','Amygdala']

drift_En_tissue_df.to_csv(results_dir+'/'+'BBB_drift_polyfit_sigEn_on_T1_std.csv', mode='a', header= False)

print("Data appended successfully in directory: /gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_manuscript_results")
