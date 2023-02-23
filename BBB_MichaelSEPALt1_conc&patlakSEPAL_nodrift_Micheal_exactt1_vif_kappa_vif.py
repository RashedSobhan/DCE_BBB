#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Oct 12 11:44:54 2022

@author: wae16sru
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Oct  3 22:46:51 2022

@author: wae16sru (Rashed Sobhan)
"""
####this code has the modification suggested by MIchael in the email dated 8/12/2022 with jupyter notebok fit_dce.ipynb
#####anything that has ################ before the modified code is RS code 
####this code was developed on 23/02/2023 to include exact VIF voxels (manually selected) with first 4 points excluded from fitting 

# this file does t1 fitting according to SEPAL 
# signal to concentration conversion was done by SEPAL too 
import os

sub_ID= input("enter subject ID: ")
sub_ID_split=sub_ID.split("_")
sub_ID_short=int(sub_ID_split[0])

dir= "/gpfs/home/wae16sru/BBB_sample_data/" + sub_ID
os.chdir(str(dir))
results_dir= "/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_manuscript_results/BBB_DCE_results"

import sys
import numpy as np
import matplotlib.pyplot as plt
sys.path.append('/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/SEPAL-master/src')
import t1_fit,dce_fit, relaxivity, signal_models, water_ex_models, aifs, pk_models

#import from Donnie

import math
import numpy as np
import nibabel as nib
import matplotlib.pyplot as plt
plt.style.use( 'ggplot' )
plt.ion()
from scipy.optimize import minimize, newton, basinhopping, least_squares

#import for excel I/O

import sys
# import os
import openpyxl
import pandas as pd
from openpyxl import Workbook

sys.path.append('/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification')

import BBB_drift_average_ROI as BBB_drift
sigEn_tissue_drift= BBB_drift.sigEn_tissue_drift

sigEn_WM_drift = BBB_drift.sigEn_WM_drift
sigEn_GM_drift = BBB_drift.sigEn_GM_drift
sigEn_CSF_drift = BBB_drift.sigEn_CSF_drift
sigEn_Thalamus_drift = BBB_drift.sigEn_Thalamus_drift
sigEn_Hippo_drift = BBB_drift.sigEn_Hippo_drift
sigEn_Amygdala_drift = BBB_drift.sigEn_Amygdala_drift

 

book_input = openpyxl.load_workbook('/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_excel_input.xlsx')
#manually create the excel input worksheet first 
sheet = book_input.active

hct_cell= sheet.cell (row= int(sub_ID_split[0])+1,column=6) 
ss_x_cell=sheet.cell (row= int(sub_ID_split[0])+1,column=3) 
ss_y_cell=sheet.cell (row= int(sub_ID_split[0])+1,column=4) 
ss_z_cell= sheet.cell (row= int(sub_ID_split[0])+1,column=5) 

hct=hct_cell.value
ss_x=ss_x_cell.value
ss_y=ss_y_cell.value
ss_z=ss_z_cell.value

## MAIN CODE LOOP_from DONNIE
# Load spoiled GRE images (in nifti format) from the working directory
T1img1 = nib.load('despot1_5_BET_MSK.nii.gz')
T1img1_data = T1img1.get_data( )
hdr_T1img1 = T1img1.header

T1img2 = nib.load('despot1_10_BET_MSK.nii.gz')
T1img2_data = T1img2.get_data( )

T1img3 = nib.load('despot1_15_BET_MSK.nii.gz')
T1img3_data = T1img3.get_data( )

B1img = nib.load('despot1_IR5_BET_MSK.nii.gz')
B1img_data = B1img.get_data( )

# Load DCE time series data from working directory
DCEimg = nib.load('dce_BET_MSK.nii.gz')
DCEimg_data = DCEimg.get_data( )



#######this is new , awe are calcaulting kapp from the kappa map created for each subject
kappaimg = nib.load(dir + '/T1-mapping_results_RS/k_fa_HIFI.nii')
kappa_data = kappaimg.get_data( )

t1_map = nib.load(dir + '/T1-mapping_results_RS/t1_HIFI.nii')
t1_data = t1_map.get_data( )

# Load parenchyma/CSF segmentation masks from working directory
#WM = nib.load('t1S_seg_2.nii.gz') # White matter
#WM_mask = WM.get_data( )
#GM = nib.load('t1S_seg_1.nii.gz') # Grey matter
#GM_mask = GM.get_data( )
#CSF = nib.load('t1S_seg_0.nii.gz') # Cerebrospinal fluid
#CSF_mask = CSF.get_data( )

WM= nib.load (dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/total_WM.nii.gz")
WM_mask = WM.get_data( )
CSF= nib.load (dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/CSF_fastsurfer_rawavg.nii.gz")
CSF_mask = CSF.get_data( )
GM= nib.load (dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/total_GM.nii.gz")
GM_mask = GM.get_data( )
hippo= nib.load(dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/total_hippo.nii.gz")
hippo_mask=hippo.get_data()
thalamus= nib.load(dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/total_thalamus.nii.gz")
thalamus_mask= thalamus.get_data()
amygdala= nib.load(dir + "/fastsurfer_output_on_T1_std/mri/ROI_RS_on_T1_std/total_amygdala.nii.gz")
amygdala_mask= amygdala.get_data()
whole_brain=nib.load(dir + "/brain_mask_FINAL.nii.gz")
whole_brain_mask=whole_brain.get_data()

# TIME SERIES SIGNAL ENHANCEMENT CALCULATIONS
#==============================================
# Calculate timings of each acquisition time point. ** USER INPUT **
acq_len = 591.5  # Acquisition duration (s) (subtracted 31.5 so pt 0 = time 0)
nPts = len( DCEimg_data[ 1, 1, 1, : ] )
###########tVec = np.arange( 0, nPts ) * acq_len / nPts

tVec= (np.arange( 0, nPts ) + 0.5) * acq_len / nPts
# The vascular input function is calculated from a voxel in the superior
# sagittal sinus. Give co-ordinates for that voxel here.

#ss_x= input("enter the VIF x: ")
#ss_y= input("enter the VIF y: ")
#ss_z= input("enter the VIF z: ")

#ss_vox= [int("ss_x",base=10), int("ss_y",base=10), int("ss_z",base=10)]
ss_vox= [ss_x,ss_y,ss_z]

# Calculate median signal intensities for each tissue type, for each time point
# Initialise arrays to hold median tissue SIs per image
med_WM = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )  # No baseline
med_GM = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )
med_CSF = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )
med_hippo = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )
med_thalamus = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )
med_amygdala = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )
# No median measurement for sagittal sinus - we use one voxel only
SS = np.zeros( len( DCEimg_data[ 1, 1, 1, : ] ) )

# No median signal intensity at baseline from DESPOT1_15deg ( different TR/TE )

# Calculate median signal per time point by looping through DCE images
for i in range( 0, len( DCEimg_data[ 1, 1, 1, : ] ) ):
    DCEtPt = DCEimg_data[ :, :, :, i ]  # Split DCE data into time points
    med_WM[ i ] = np.median( DCEtPt[ np.where( WM_mask > 0 ) ] )
    med_GM[ i ] = np.median( DCEtPt[ np.where(( WM_mask == 0) & ( CSF_mask == 0) & (whole_brain_mask > 0)) ] )
    med_CSF[ i ] = np.median( DCEtPt[ np.where( CSF_mask > 0 ) ] )
    med_hippo [i]= np.median( DCEtPt[ np.where( hippo_mask > 0 ) ] )
    med_thalamus [i]= np.median( DCEtPt[ np.where( thalamus_mask > 0 ) ] )
    med_amygdala [i]= np.median( DCEtPt[ np.where( amygdala_mask > 0 ) ] )
    SS[ i ] = DCEtPt[ ss_vox[ 0 ], ss_vox[ 1 ], ss_vox[ 2 ] ]

# Convert to signal enhancement from baseline (ratio, use % for plotting only)
sigEn_WM = (( med_WM - med_WM[ 0 ] ) / med_WM[ 0 ]) 
sigEn_GM = (( med_GM - med_GM[ 0 ] ) / med_GM[ 0 ]) 
sigEn_CSF = (( med_CSF - med_CSF[ 0 ] ) / med_CSF[ 0 ]) 
sigEn_hippo = (( med_hippo - med_hippo[ 0 ] ) / med_hippo[ 0 ]) 
sigEn_thalamus = (( med_thalamus - med_thalamus[ 0 ] ) / med_thalamus[ 0 ])
sigEn_amygdala = (( med_amygdala  - med_amygdala[ 0 ] ) / med_amygdala[ 0 ]) 
#????more regions to add here 
sigEn_SS = ( SS - SS[ 0 ] ) / SS[ 0 ]


sigEn_tissue=[sigEn_WM, sigEn_GM, sigEn_CSF, sigEn_hippo, sigEn_thalamus, sigEn_amygdala] # ???need modification after more tissue regions 
# changes to adapt SEPAL 

# making a csv file of all signal enhancement from each subject, needed to draw average signal drift from each cohort. 

sigEn_tissue_df = pd.DataFrame (sigEn_tissue)

sigEn_tissue_df.index = ['WM','GM','CSF','Hippo','Thalamus','Amygdala']

sigEn_tissue_df.to_csv(results_dir+'/'+f'{sub_ID}_sigEn_allROI.csv', mode='a', header= False)



t = tVec #variable name change to use SEPALsigEn_WM
s_vif = sigEn_SS 
########t1_vif = 1.584 
t1_vif= t1_data[ss_vox[0], ss_vox[1], ss_vox[2]] 

#######k_vif = 1 # assuming no B1 distortion for vein 
k_vif= kappa_data[ss_vox[0], ss_vox[1], ss_vox[2]] 
#hct = 0.423

'''DCam. The first few participants have different TR/TE for DCE-MRI. Will use
t = 0 DCE timepoint as baseline instead ...'''

# Plotting code - shows signal intensity time-curves for WM, GM, CSF, sigEn_SS SS
fig = plt.figure( )
fig.suptitle( 'Signal Enhancement Curves', fontsize = 14 )
ax1 = plt.gca( )

plt1, = plt.plot( tVec, sigEn_WM * 100, '#56B4E9', label = 'White Matter', \
linewidth = 2 )
plt2, = plt.plot( tVec, sigEn_GM * 100, '#0072B2', label = 'Grey Matter' , \
linewidth = 2 )
plt3, = plt.plot( tVec, sigEn_hippo * 100, '#CC79A7', label = 'Hippocampus' , \
linewidth = 2 )
plt4, = plt.plot( tVec, sigEn_thalamus * 100, '#00FFFF', label = 'Thalamus' , \
linewidth = 2 )
plt5, = plt.plot( tVec, sigEn_amygdala * 100, '#8A2BE2', label = 'Amygdala' , \
linewidth = 2 )
plt6, = plt.plot( tVec, sigEn_CSF * 100, '#8B7355', label ='CSF' , \
linewidth = 2 )
#????more regions to add here 

ax1.set_xlabel( 'Time (s)' )
ax1.set_ylabel( 'Signal Enhancement (%)' )
ax1.grid(None)
# Put SS on similar scale
ax2 = ax1.twinx( )
plt7, = plt.plot( tVec, sigEn_SS * 100, '#D55E00', label = 'Sag. Sinus' , \
linewidth = 2 )
ax2.set_ylabel( 'SSS Signal Enhancement (%)', color = '#D55E00' )
ax2.tick_params( 'y', colors = '#D55E00' )
ax2.set_xlim( [ 0, np.max( tVec ) + 20 ] )
ax2.grid(None)

# plt.legend( handles = [ plt1, plt2, plt3, plt4,plt5, plt6, plt7 ], loc = 'lower center', mode = "expand", ncol = 4,  frameon = False ) 
plt.legend( handles = [ plt1, plt2, plt3, plt4, plt5, plt7 ], bbox_to_anchor=(1, -0.15), loc="lower right",bbox_transform=fig.transFigure, ncol=4) 

#plt.show( )
#????more regions to add here 

plt.savefig(results_dir+f'/{sub_ID}_RS_sigEn_plot_nodrift_on_T1_std.png', format = 'png', bbox_inches = 'tight', dpi = 500)
plt.savefig(dir+f'/{sub_ID}_RS_sigEn_plot_nodrift_on_T1_std.png', format = 'png', bbox_inches = 'tight', dpi = 500)




#parmaeters from Donnie
alpha_p = [ 5, 10, 15 ] # List of prescribed flip angles for spGRE (deg).
alpha_P = 5             # List of prescribed flip angles for IR-spGRE (deg).
TI_IR = 0.45            # Inversion time (s)
TR = 0.005568           # Repetition time of spoiled GRE sequence (s)
TE = 0.001508           # Echo time of spoiled GRE sequence (s)
TR_IR = 0.8             # Time between successive inversion pulses (s)
TD_IR=0
N= 96
PECentre=0

alpha_pRad = [ 0 ] * len( alpha_p ) # Initialise list of flips in rads

for fa in range( 0, len( alpha_p ) ):
    alpha_pRad[ fa ] = alpha_p[ fa ] * math.pi / 180


alpha_PRad = alpha_P * math.pi / 180    # IR flip angle in rads

medWM_sigSPGR = [ np.median( T1img1_data[ np.where( WM_mask > 0 ) ] ), \
np.median( T1img2_data[ np.where( WM_mask > 0 ) ] ), \
med_WM[ 0 ] ] # Arrange WM SPGR data and redefine med_WM[ 0 ] for clarity
medWM_sigIR = np.median( B1img_data[ np.where( WM_mask > 0 ) ] )

medGM_sigSPGR = [ np.median( T1img1_data[ np.where( ( WM_mask == 0) & ( CSF_mask == 0) & (whole_brain_mask > 0))   ] ), \
np.median( T1img2_data[ np.where( ( WM_mask == 0) & ( CSF_mask == 0) & (whole_brain_mask > 0) ) ] ), \
med_GM[ 0 ] ] # Arrange GM SPGR data and redefine med_GM[ 0 ] for clarity
medGM_sigIR = np.median( B1img_data[ np.where( ( WM_mask == 0) & ( CSF_mask == 0) & (whole_brain_mask > 0)) ] )

medCSF_sigSPGR = [ np.median( T1img1_data[ np.where( CSF_mask > 0 ) ] ), \
np.median( T1img2_data[ np.where( CSF_mask > 0 ) ] ), \
med_CSF[ 0 ] ] # Arrange CSF SPGR data and redefine med_CSF[ 0 ] for clarity
medCSF_sigIR = np.median( B1img_data[ np.where( CSF_mask > 0 ) ] )

#????more regions to add here 

medhippo_sigSPGR = [ np.median( T1img1_data[ np.where( hippo_mask > 0 ) ] ), \
np.median( T1img2_data[ np.where( hippo_mask > 0 ) ] ), \
med_hippo[ 0 ] ] # Arrange CSF SPGR data and redefine med_CSF[ 0 ] for clarity
medhippo_sigIR = np.median( B1img_data[ np.where( hippo_mask > 0 ) ] )


medthalamus_sigSPGR = [ np.median( T1img1_data[ np.where ( thalamus_mask > 0 ) ] ), \
np.median( T1img2_data[ np.where( thalamus_mask > 0 ) ] ), \
med_thalamus[ 0 ] ] # Arrange CSF SPGR data and redefine med_CSF[ 0 ] for clarity
medthalamus_sigIR = np.median( B1img_data[ np.where( thalamus_mask > 0 ) ] )

medamygdala_sigSPGR = [ np.median( T1img1_data[ np.where(amygdala_mask > 0 ) ] ), \
np.median( T1img2_data[ np.where(amygdala_mask > 0 ) ] ), \
med_amygdala[ 0 ] ] # Arrange CSF SPGR data and redefine med_CSF[ 0 ] for clarity
medamygdala_sigIR = np.median( B1img_data[ np.where( amygdala_mask > 0 ) ] )



#DESPOT_HIFI method from Michael (demo_fit_t1)

# s = np.array([249, 585, 413, 604, 445]) # signal
# s=medWM_sigSPGR

s_SPGR = np.vstack ((medWM_sigSPGR,  medGM_sigSPGR, medCSF_sigSPGR, medhippo_sigSPGR, medthalamus_sigSPGR , medamygdala_sigSPGR ))
s_IR = np.vstack ((medWM_sigIR, medGM_sigIR, medCSF_sigIR, medhippo_sigIR, medthalamus_sigIR, medamygdala_sigIR ))# signal
# s.insert (0, medWM_sigIR)
s_tissue = np.concatenate((s_IR,s_SPGR), axis = 1)
esp = np.array([5.564e-3, 5.568e-3, 5.568e-3, 5.568e-3]) # echo spacing (IR-SPGR) or TR (SPGR) ?????
# esp = np.array([1.5e-3, 5.568e-3, 5.568e-3, 5.568e-3]) 
ti = np.array([0.45, np.nan, np.nan, np.nan]) # delay after inversion pulse
n = np.array([96, np.nan, np.nan, np.nan]) # number of readout pulses (IR-SPGR only)
b = np.array([5, 5, 10, 15]) # excitation flip angle
td = np.array([0, np.nan, np.nan, np.nan]) # delay between end of readout train and next inversion pulse (IR-SPGR only)
centre = np.array([0, np.nan, np.nan, np.nan]) # time when centre of k-space is acquired (expressed as fraction of readout pulse train length; IR-SPGR only)

# s0_tissue, t1_tissue, k_fa_tissue= np.zeros((s_tissue.shape[0],1))
# s_fit_tissue=np.zeros(s_tissue.shape)CSF_fastsurfer_rawavg.nii
s0 = np.zeros((s_tissue.shape[0],1)) 
t1 = np.zeros((s_tissue.shape[0],1))
k_fa = np.zeros((s_tissue.shape[0],1))

s_fit = np.zeros(s_tissue.shape)
s = np.zeros(s_tissue.shape)
for i in range(0, s_tissue.shape[0]): 
    s = s_tissue[i,:]
    t1_calculator = t1_fit.HIFI(esp, ti, n, b, td, centre)

    s0[i], t1[i], k_fa[i], s_fit[i] = t1_calculator.proc(s)
    

#the follwoing bit needs correction 
#for i in range(0, len(s0)):
    
    #print(f"Fitted values: s0 = {s0[i]:.1f}, t1 = {t1[i]:.3f} s, k_fa = {k_fa[i]:.3f}" for i in range(0, s_tissue.shape[0]))
    #plt.plot(np.arange(4), s_fit[i], '-', label='model')
    #plt.plot(np.arange(4), s0, 'o', label='signal')
    #plt.xlabel('acquisition number')
    #plt.ylabel('signal');
    #plt.legend();
# correct that bit to plot ???????RS

T1_WM = t1[0]
RHO_WM = s0[0]
KAPPA_WM = k_fa[0]

T1_GM = t1[1]
RHO_GM = s0[1]
KAPPA_GM = k_fa[1]

T1_CSF = t1[2]
RHO_CSF = s0[2]
KAPPA_CSF = k_fa[2]

T1_hippo=t1[3]
RHO_CSF = s0[3]
KAPPA_CSF = k_fa[3]

T1_thalamus=t1[4]
RHO_thalamus= s0[4]
KAPPA_thalamus = k_fa[4]

T1_amygdala=t1[5]
RHO_amygdala = s0[5]
KAPPA_amygdala = k_fa[5]



# First define the relationship between concentration and relaxation rate:
# c_to_r_model = relaxivity.CRLinear(r1=5.0, r2=7.1)
######c_to_r_model = relaxivity.CRLinear(r1=4.5, r2=6.3); used this in the loop_ROi_nodrift.py
c_to_r_model = relaxivity.CRLinear(r1=5.0, r2=0) 

# and the relationship between relaxation rate and signal:
# signal_model = signal_models.SPGR(tr=5.7e-3, fa=15, te=1.7e-3)
signal_model = signal_models.SPGR(tr=6.4e-3, fa=15, te=2.5e-3)


# Now create an EnhToConc object and use the proc method to get concentration
#double check which one to use for EnhToConc or EnhToConcSPGR ???????????????????????????, teh data was obtained with 

#e_to_c = dce_fit.EnhToConc(c_to_r_model, signal_model,C_min=-0.5, C_max=30, n_samples=1000)
###### e_to_c = dce_fit.EnhToConc(c_to_r_model, signal_model)
#e_to_c = dce_fit.EnhToConcSPGR(tr=6.4e-3, fa=15, r1=4.5)
e_to_c = dce_fit.EnhToConcSPGR(tr=6.3793e-3, fa=15, r1=5.0)  # alternative method - faster, assumes r2=0
ROI=['WM', 'GM', 'CSF', 'Hippo', 'Thalamus', 'Amygdala']

#tissue_type= int (input (" enter 0 for WM, 1 for GM, 2 for CSF, 3 for hippo, 4 for thalamus, 5 for amygdala, so on: "), base=10)
for tissue_type in range(6):
    #for 5 tissue types range has to be 6
    # C_t = e_to_c.proc(enh_tissue, t1_tissue, k_tissue)
    C_t = e_to_c.proc(100* sigEn_tissue[tissue_type], t1[tissue_type], k_fa [tissue_type])  
    
    # multiplication by 100 gives values that Michael expected  , e.g. GM 
    #C_t = e_to_c.proc(sigEn_tissue[tissue_type], t1[tissue_type], k_fa [tissue_type])
    # c_p_vif = e_to_c.proc(enh_vif, t1_vif, k_vif) / (1-hct)
    
    c_p_vif = e_to_c.proc(100* sigEn_SS, t1_vif, k_vif) / (1.0-hct)
    #c_p_vif = e_to_c.proc(sigEn_SS, t1_vif, k_vif) / (1.0-hct)
    
    fig, ax = plt.subplots(1,2)
    ax[0].plot(t, C_t, '.-', label= ROI[tissue_type]+' conc. (mM)')
    ax[1].plot(t, c_p_vif, '.-', label='VIF plasma conc. (mM)')
    ax[1].set_xlabel('time (s)');
    [a.legend() for a in ax.flatten()];
    
    #plt.savefig(results_dir+f"/{sub_ID}_RS_conc_VIF_"+ ROI[tissue_type] + '_nodrift_on_T1_std_enh_timed100_Michael_exact_t1vif_kappavif.png', format = 'png', bbox_inches = 'tight', dpi = 500)
    plt.savefig(dir+f"/{sub_ID}_RS_conc_VIF_"+ ROI[tissue_type] + '_nodrift_on_T1_std_enh_timed100_Michael_exact_t1vif_kappavif.png', format = 'png', bbox_inches = 'tight', dpi = 500)


   


    # First we need to create an AIF object based on the calculate VIF concentrations:
    
    aif = aifs.PatientSpecific(t, c_p_vif)
    
    # ...and a PKModel object:
    
    pk_model = pk_models.Patlak(t, aif, bounds=((-1,-0.001),(1,1)))
  
    # Finally, we create a ConcToPKP object and use the proc method to fit the concentration data:
    
    # weights = np.concatenate([np.zeros(7), np.ones(25)]) # exclude first few points from fit
    #############weights = np.concatenate([np.zeros(5), np.ones(15)]) # exclude first few points from fit
    
    weights = np.concatenate([np.zeros(4), np.ones(16)])
    pkp_0 = [{'vp': 0.2, 'ps': 1e-4}] # starting parameters (multiple starting points can be specified if required)
    #############conc_to_pkp = dce_fit.ConcToPKP(pk_model, pkp_0, weights)

    conc_to_pkp = dce_fit.PatlakLinear(t, aif, upsample_factor=10, include=weights)
    vp, ps, C_t_fit = conc_to_pkp.proc(C_t)
    
    fig = plt.figure( )
    plt.plot(t, C_t, '.', label= ROI[tissue_type] +' conc. (mM)')
    plt.plot(t, C_t_fit, '-', label='model fit (mM)')
    plt.legend();
    #plt.savefig(results_dir+'/'+f'{sub_ID}'+'_RS_model_fit_'+ROI[tissue_type]+'_nodrift_on_T1_std_Michael_exact_t1vif_kappavif.png', format = 'png', bbox_inches = 'tight', dpi = 500)
    plt.savefig(dir+'/'+f'{sub_ID}'+'_RS_model_fit_'+ROI[tissue_type]+'_nodrift_on_T1_std_Michael_exact_t1vif_kappavif.png', format = 'png', bbox_inches = 'tight', dpi = 500)

    
    print(f"vp = {vp:.4f}, ps = {ps:.6f}")
    
    #book_result= Workbook()
    book_result= openpyxl.load_workbook('/gpfs/home/wae16sru/BBB_sample_data/BBB_Michael_Code_modification/BBB_results_ADA_no_drift_on_T1_std_enh_timed100_Michael_exact_t1vif_kappavif.xlsx')
    #book_result= Workbook()
    sheet_result=book_result.active
    sheet_result.cell(row=int(sub_ID_split[0])+1, column = 5*(tissue_type)+7).value= vp
    sheet_result.cell(row=int(sub_ID_split[0])+1, column= 5*(tissue_type)+8).value= ps
    sheet_result.cell(row=int(sub_ID_split[0])+1, column = 5*(tissue_type)+9).value= float(t1[tissue_type])
    sheet_result.cell(row=int(sub_ID_split[0])+1, column = 5*(tissue_type)+10).value= float(s0[tissue_type])
    sheet_result.cell(row=int(sub_ID_split[0])+1, column = 5*(tissue_type)+11).value= float(k_fa[tissue_type])
    
    book_result.save(results_dir+"BBB_results_ADA_no_drift_on_T1_std_enh_timed100_Michael_exact_t1vif_kappavif_exactVIF_4pointsexcluded.xlsx")
    
    df = pd.DataFrame(pd.read_excel(results_dir+"BBB_results_ADA_no_drift_on_T1_std_enh_timed100_Michael_exact_t1vif_kappavif_exactVIF_4pointsexcluded.xlsx"))
    #df= pd.DataFrame(pd.read_excel("C:\\Users\\wae16sru\\OneDrive - University of East Anglia\\BBB_results\\BBB_results.xlsx"))
    
    print(df)

# for presentation purpose 

roi_names = ('WM', 'GM', 'CSF', 'Hippocampus', 'Thalamus', 'Amygdala')

# plot enhancements
plt.figure(figsize=(15,8))
for idx, roi in enumerate(roi_names):
    plt.subplot(2,3,idx+1)
    plt.plot(t, sigEn_tissue[idx], 'ko-')
    plt.title(roi+"_Enhancement")

plt.figure()
plt.plot(t, sigEn_SS, 'r-')
plt.title('VIF');

plt.savefig(results_dir+'/'+f'{sub_ID}'+'_sigEn'+'_nodrift_on_T1_std_Michael_exact_t1vif_kappavif.png', format = 'png', bbox_inches = 'tight', dpi = 500)

C_t_tissue = [ e_to_c.proc(100*sigEn_tissue[idx], t1[idx], k_fa[idx]) for idx in range(len(roi_names)) ]
c_p_vif = e_to_c.proc(sigEn_SS, t1_vif, k_vif) / (1-hct)



# plot concentrations
plt.figure(figsize=(15,8))
for idx, roi in enumerate(roi_names):
    plt.subplot(2,3,idx+1)
    plt.plot(t, C_t_tissue[idx], 'ko-')
    plt.title(roi+"_Conc")

plt.figure()
plt.plot(t, c_p_vif, 'r-')
plt.title('VIF');
plt.savefig(results_dir+'/'+f'{sub_ID}'+'_sigEn_to_Conc'+'_nodrift_on_T1_std_Michael_exact_t1vif_kappavif.png', format = 'png', bbox_inches = 'tight', dpi = 500)


vp, ps, C_t_fit = list(zip(*
                   [ conc_to_pkp.proc(conc) for conc in C_t_tissue ]
                  ))
#plot fits 
plt.figure(figsize=(15,10))
for idx, roi in enumerate(roi_names):
    plt.subplot(2,3,idx+1)
    plt.plot(t[:3], C_t_tissue[idx][:3], 'ko', fillstyle='none')
    plt.plot(t[4:], C_t_tissue[idx][4:], 'ko')
    plt.plot(t, C_t_fit[idx], 'k:')
    plt.title(f'{roi}'+'_fit\nvp = {vp[idx]:.4f}, ps = {ps[idx]:.6f}')


plt.savefig(results_dir+'/'+f'{sub_ID}'+'_Conc_fit'+'_nodrift_on_T1_std_Michael_exact_t1vif_kappavif.png', format = 'png', bbox_inches = 'tight', dpi = 500)

